import os
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.metrics import confusion_matrix, f1_score, precision_score, recall_score
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows


# 创建文件夹
def create_directory(path):
    if not os.path.exists(path):
        os.makedirs(path)


# 绘制训练和验证集的准确率和损失图像
def plot_metrics(metrics, output_path, labels):
    create_directory(output_path)  # 创建输出目录
    epochs = range(1, metrics['Epoch'] + 2)

    # 绘制训练和验证集的准确率
    plt.figure(figsize=(8, 5))
    plt.plot(epochs, metrics['epoch_accuracy_list'], 'b', label='Training accuracy')
    plt.plot(epochs, metrics['test_accuracy_list'], 'r', label='Validation accuracy')
    plt.xlabel('Epochs')
    plt.ylabel('Accuracy')
    plt.title('Training and Validation Accuracy')
    plt.legend()
    plt.savefig(f'{output_path}/accuracy_plot.png')
    plt.close()

    # 绘制训练和验证集的损失
    plt.figure(figsize=(8, 5))
    plt.plot(epochs, metrics['epoch_loss_list'], 'b', label='Training loss')
    plt.plot(epochs, metrics['test_loss_list'], 'r', label='Validation loss')
    plt.xlabel('Epochs')
    plt.ylabel('Loss')
    plt.title('Training and Validation Loss')
    plt.legend()
    plt.savefig(f'{output_path}/loss_plot.png')
    plt.close()

    # 绘制混淆矩阵
    cm = confusion_matrix(metrics['all_labels'], metrics['all_preds'])
    plt.figure(figsize=(10, 7))
    sns.heatmap(cm, annot=True, fmt='d', cmap='Blues', xticklabels=labels, yticklabels=labels)
    plt.xlabel('Predicted')
    plt.ylabel('True')
    plt.title('Confusion Matrix')
    plt.savefig(f'{output_path}/confusion_matrix.png')
    plt.close()


# 计算并打印模型设定参数和性能参数
def get_model_parameters_and_performance(model_name, metrics):
    model_params = {
        "Model Name": model_name,
        "Training epochs": metrics['Epoch'],
        "Batch size": metrics['batch_size'],
        "Learning rate": metrics['learning_rate'],
        "Model parameters": metrics['Model Parameters'],
        "Total training time": metrics['Total Training Time'],
        "Prediction time for one image": metrics['Prediction Time for one image'],
        "Final Training Accuracy": metrics['epoch_accuracy_list'][-1],
        "Final Training Loss": metrics['epoch_loss_list'][-1],
        "Final Validation Accuracy": metrics['test_accuracy_list'][-1],
        "Final Validation Loss": metrics['test_loss_list'][-1],
        "F1 Score": f1_score(metrics['all_labels'], metrics['all_preds'], average='weighted'),
        "Precision": precision_score(metrics['all_labels'], metrics['all_preds'], average='weighted'),
        "Recall": recall_score(metrics['all_labels'], metrics['all_preds'], average='weighted')
    }
    return model_params


# 将参数和图像写入Excel文件
def write_to_excel(metrics, output_path, model_name):
    create_directory(output_path)  # 确保输出目录存在

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Parameters"

    # 写入参数
    model_params = get_model_parameters_and_performance(model_name, metrics)
    for i, (key, value) in enumerate(model_params.items(), start=1):
        ws1.cell(row=i, column=1, value=key)
        ws1.cell(row=i, column=2, value=value)

    # 创建第二个sheet
    ws2 = wb.create_sheet(title="Metrics and Plots")
    df = pd.DataFrame({
        'Epoch': range(1, metrics['Epoch'] + 2),
        'Train Accuracy': metrics['epoch_accuracy_list'],
        'Train Loss': metrics['epoch_loss_list'],
        'Validation Accuracy': metrics['test_accuracy_list'],
        'Validation Loss': metrics['test_loss_list']
    })
    for r in dataframe_to_rows(df, index=False, header=True):
        ws2.append(r)

    # 添加图像到第二个sheet
    img1_path = f'{output_path}/accuracy_plot.png'
    img2_path = f'{output_path}/loss_plot.png'
    img3_path = f'{output_path}/confusion_matrix.png'

    img1 = Image(img1_path)
    img2 = Image(img2_path)
    img3 = Image(img3_path)

    ws2.add_image(img1, 'F2')
    ws2.add_image(img2, 'F28')
    ws2.add_image(img3, 'F54')

    # 保存Excel文件
    excel_path = f"{output_path}/model_metrics.xlsx"
    wb.save(excel_path)
    print(f"Excel file saved at {excel_path}")


def parameters_save(filepath, metrics, model_name, labels):
    plot_metrics(metrics, filepath, labels)
    write_to_excel(metrics, filepath, model_name)


# metrics = {
#     "Epoch": 10,
#     "batch_size": 32,
#     "learning_rate": 0.001,
#     "Total Training Time": '12.345 min',
#     "Model Parameters": '1.23M',
#     "Prediction Time for one image": '1.234ms',
#     "epoch_accuracy_list": [0.5, 0.6, 0.65, 0.7, 0.75, 0.8, 0.85, 0.86, 0.88, 0.9],
#     "epoch_loss_list": [1.2, 1.0, 0.9, 0.85, 0.8, 0.75, 0.7, 0.68, 0.65, 0.6],
#     "test_accuracy_list": [0.48, 0.58, 0.63, 0.68, 0.73, 0.78, 0.82, 0.83, 0.85, 0.88],
#     "test_loss_list": [1.25, 1.05, 0.95, 0.88, 0.83, 0.77, 0.72, 0.70, 0.68, 0.65],
#     "all_labels": [0, 1, 0, 1, 0, 1, 0, 1, 0, 1],  # 示例数据
#     "all_preds": [0, 1, 0, 1, 1, 1, 0, 1, 0, 0]  # 示例数据
# }

# parameters_save('./parameters_save', metrics)
